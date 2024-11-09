import openpyxl
import re

# 加载第一个和第二个 .xlsx 文件
file1_path = "D:\\Works\\files\\副本21天单词打卡记录(1).xlsx"  # 第一个文件的路径
file2_path = "D:\\Works\\files\\excel_9192587_圈子成员_9192587_1731153224339_512d5a.xlsx"  # 第二个文件的路径
file1 = openpyxl.load_workbook(file1_path)
file2 = openpyxl.load_workbook(file2_path)

# 选择活动工作表
sheet1 = file1.active
sheet2 = file2.active

day = int(input("请问您想将数据写入第几列（11/9对应D列，即第4列，因此输入4，以此类推）："))

# 获取第一个文件的姓名列表（A列，跳过表头）
names_in_file1 = [sheet1.cell(row=row, column=1).value for row in range(2, sheet1.max_row + 1)]

# 定义一个函数，在第二个文件中的A列和C列中查找姓名
def find_total_days_in_file2(name):
    pattern = re.compile(re.escape(name))
    for row in range(2, sheet2.max_row + 1):
        col_a_value = sheet2.cell(row=row, column=1).value
        col_c_value = sheet2.cell(row=row, column=3).value

        # 搜索姓名的正则表达式匹配
        if (col_a_value and pattern.search(str(col_a_value))) or (col_c_value and pattern.search(str(col_c_value))):
            # 返回对应行的F列值（总打卡天数）
            return sheet2.cell(row=row, column=6).value
    return None  # 如果未找到姓名

# 循环遍历第一个文件的姓名列表，在第二个文件中查找总打卡天数并写入D列（以11/9为例）
for idx, name in enumerate(names_in_file1, start=2):  # 从第2行开始，跳过表头
    total_days = find_total_days_in_file2(name)
    if total_days is not None:
        sheet1.cell(row=idx, column=day, value=total_days)  # 将总打卡天数写入第一个文件的D列

# 保存第一个文件的更改
updated_file1_path = "file1_updated.xlsx"
file1.save(updated_file1_path)
print(f"数据已成功写入 {updated_file1_path}")
